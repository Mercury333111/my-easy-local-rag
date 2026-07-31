"""
upload.py - Document upload and management for the RAG system.

Usage:
    python upload.py                    # Launch GUI file picker
    python upload.py <file>             # Upload a specific file
    python upload.py --list             # List uploaded documents
    python upload.py --remove <name>    # Remove a document by name
    python upload.py --clear            # Remove all documents
"""

import sys
sys.stdin.reconfigure(encoding='utf-8')
sys.stdout.reconfigure(encoding='utf-8')

import os
import re
import json
import csv
import argparse
from io import StringIO

from core import (
    load_config, load_vault, save_vault, chunk_text,
    load_registry, save_registry, registry_sources,
    remove_source_from_vault, clear_vault, Colors as C,
)


# ===== File readers =====
# Each reader returns a plain text string from the file.

def read_pdf(file_path):
    """Read text from a PDF file."""
    import PyPDF2
    with open(file_path, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + " "
    return text


def read_txt(file_path):
    """Read text from a plain text file."""
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()


def read_json(file_path):
    """Read text from a JSON file (flattened)."""
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return json.dumps(data, ensure_ascii=False, indent=2)


def read_markdown(file_path):
    """Read text from a Markdown file (strip formatting)."""
    with open(file_path, 'r', encoding='utf-8') as f:
        text = f.read()
    # Remove markdown syntax for cleaner chunks
    text = re.sub(r'```[\s\S]*?```', '', text)  # code blocks
    text = re.sub(r'`[^`]+`', '', text)           # inline code
    text = re.sub(r'#+\s*', '', text)             # headers
    text = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', text)  # links
    text = re.sub(r'[*_~]{1,3}', '', text)        # bold/italic/strikethrough
    text = re.sub(r'!\[[^\]]*\]\([^)]+\)', '', text)  # images
    return text


def read_docx(file_path):
    """Read text from a Word document."""
    from docx import Document
    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def read_html(file_path):
    """Read text from an HTML file."""
    from bs4 import BeautifulSoup
    with open(file_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f.read(), 'lxml')
    # Remove script and style elements
    for tag in soup(['script', 'style', 'nav', 'footer', 'header']):
        tag.decompose()
    return soup.get_text(separator=' ', strip=True)


def read_csv(file_path):
    """Read text from a CSV file."""
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return ""
    # Use first row as headers
    headers = rows[0]
    lines = []
    for row in rows[1:]:
        pairs = [f"{h}: {v}" for h, v in zip(headers, row) if v.strip()]
        if pairs:
            lines.append(" | ".join(pairs))
    return "\n".join(lines)


def read_xlsx(file_path):
    """Read text from an Excel file."""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    all_text = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h else "" for h in rows[0]]
        for row in rows[1:]:
            pairs = [f"{h}: {v}" for h, v in zip(headers, row) if v is not None and str(v).strip()]
            if pairs:
                all_text.append(" | ".join(pairs))
    wb.close()
    return "\n".join(all_text)


# ===== Format dispatch =====

READERS = {
    '.pdf': read_pdf,
    '.txt': read_txt,
    '.json': read_json,
    '.md': read_markdown,
    '.docx': read_docx,
    '.html': read_html,
    '.htm': read_html,
    '.csv': read_csv,
    '.xlsx': read_xlsx,
}


def upload_file(file_path, config):
    """Upload a single file to the vault."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in READERS:
        print(f"{C.RED}Unsupported format: {ext}{C.RESET}")
        print(f"Supported: {', '.join(READERS.keys())}")
        return False

    vault_file = config["vault_file"]
    registry_file = config["registry_file"]
    chunk_size = config.get("chunk_size", 500)
    chunk_overlap = config.get("chunk_overlap", 100)
    source_name = os.path.basename(file_path)

    # Check if already uploaded
    existing_sources = registry_sources(registry_file)
    if source_name in existing_sources:
        print(f"{C.YELLOW}'{source_name}' already exists. Removing old version first...{C.RESET}")
        remove_source_from_vault(vault_file, registry_file, source_name)

    # Read file
    print(f"{C.GREEN}Reading {ext} file...{C.RESET}")
    try:
        reader = READERS[ext]
        text = reader(file_path)
    except Exception as e:
        print(f"{C.RED}Error reading file: {e}{C.RESET}")
        return False

    if not text.strip():
        print(f"{C.RED}File is empty or could not extract text.{C.RESET}")
        return False

    # Chunk text
    chunks = chunk_text(text, chunk_size=chunk_size, overlap=chunk_overlap)
    if not chunks:
        print(f"{C.RED}No chunks produced from file.{C.RESET}")
        return False

    # Append to vault and registry
    vault_lines = load_vault(vault_file)
    registry = load_registry(registry_file)
    start_index = len(vault_lines)

    for i, chunk in enumerate(chunks):
        vault_lines.append(chunk)
        registry.append({
            "source": source_name,
            "chunk_index": i,
            "text_preview": chunk[:80],
        })

    save_vault(vault_file, vault_lines)
    save_registry(registry_file, registry)

    print(f"{C.GREEN}✓ Uploaded '{source_name}': {len(chunks)} chunks added (total: {len(vault_lines)}){C.RESET}")
    return True


def list_documents(config):
    """List all uploaded documents."""
    registry_file = config["registry_file"]
    registry = load_registry(registry_file)

    if not registry:
        print(f"{C.YELLOW}No documents uploaded yet.{C.RESET}")
        return

    # Group by source
    sources = {}
    for entry in registry:
        src = entry.get("source", "unknown")
        if src not in sources:
            sources[src] = 0
        sources[src] += 1

    print(f"{C.GREEN}Uploaded documents ({len(sources)} files, {len(registry)} chunks total):{C.RESET}")
    print("-" * 50)
    for name, count in sources.items():
        print(f"  {C.CYAN}{name}{C.RESET}  ({count} chunks)")


def remove_document(source_name, config):
    """Remove a specific document."""
    vault_file = config["vault_file"]
    registry_file = config["registry_file"]

    existing = registry_sources(registry_file)
    if source_name not in existing:
        print(f"{C.RED}'{source_name}' not found.{C.RESET}")
        print(f"Available: {', '.join(existing)}")
        return

    removed, remaining = remove_source_from_vault(vault_file, registry_file, source_name)
    print(f"{C.GREEN}✓ Removed '{source_name}': {removed} chunks deleted, {remaining} remaining.{C.RESET}")


def clear_all_documents(config):
    """Clear all documents from the vault."""
    vault_file = config["vault_file"]
    registry_file = config["registry_file"]
    cache_file = config["embeddings_cache_file"]

    clear_vault(vault_file, registry_file)

    # Also clear embedding cache
    if os.path.exists(cache_file):
        os.remove(cache_file)

    print(f"{C.GREEN}✓ All documents cleared.{C.RESET}")


# ===== GUI mode (original behavior) =====

def launch_gui(config):
    """Launch the Tkinter file picker GUI."""
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.title("RAG Document Upload")
    root.geometry("400x300")

    status_var = tk.StringVar(value="Select a file to upload")
    status_label = tk.Label(root, textvariable=status_var, wraplength=380)
    status_label.pack(pady=10)

    def do_upload():
        ext_str = " ".join(f"*{ext}" for ext in READERS)
        file_path = filedialog.askopenfilename(
            title="Select document",
            filetypes=[("All Supported", ext_str), ("All Files", "*.*")]
        )
        if file_path:
            success = upload_file(file_path, config)
            if success:
                status_var.set(f"✓ Uploaded: {os.path.basename(file_path)}")
            else:
                status_var.set(f"✗ Failed to upload")

    def do_list():
        list_documents(config)

    def do_clear():
        if messagebox.askyesno("Confirm", "Remove all documents?"):
            clear_all_documents(config)
            status_var.set("All documents cleared")

    tk.Button(root, text="Upload Document", command=do_upload, width=25, height=2).pack(pady=5)
    tk.Button(root, text="List Documents", command=do_list, width=25).pack(pady=5)
    tk.Button(root, text="Clear All", command=do_clear, width=25).pack(pady=5)

    root.mainloop()


# ===== Main =====

def main():
    config = load_config()

    parser = argparse.ArgumentParser(description="RAG Document Upload & Management")
    parser.add_argument("file", nargs="?", help="Path to file to upload")
    parser.add_argument("--list", action="store_true", help="List uploaded documents")
    parser.add_argument("--remove", metavar="NAME", help="Remove a document by name")
    parser.add_argument("--clear", action="store_true", help="Remove all documents")
    args = parser.parse_args()

    if args.list:
        list_documents(config)
    elif args.remove:
        remove_document(args.remove, config)
    elif args.clear:
        clear_all_documents(config)
    elif args.file:
        upload_file(args.file, config)
    else:
        # No arguments: launch GUI
        launch_gui(config)


if __name__ == "__main__":
    main()
